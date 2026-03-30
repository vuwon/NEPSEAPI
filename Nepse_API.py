

## 1. Import the Nepse Library

from nepse import Nepse
import json

# Initialize the Nepse object
nepse = Nepse()

# Set TLS verification to False (temporary fix for SSL certificate issue)
nepse.setTLSVerification(False)

print("Nepse API initialized successfully!")



# Save all transactions of today and selected symbols in a single combined Excel file
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import re

try:
    today = datetime.now()
    today_str = today.strftime('%Y-%m-%d')
    today_excel_date = today.date()
    print(f"Fetching all floorsheet data for today ({today_str}) (this may take a moment)...")
    floorsheet = nepse.getFloorSheet(show_progress=False)
    print(f"Total transactions for today: {len(floorsheet)}")

    # Define your fixed list of stock symbols here
    selected_symbols = [
        'BANDIPUR', 'BHCL', 'BUNGAL', 'CREST', 'DHEL', 'SAGAR', 'HIMSTAR', 'JHAPA', 'NMIC', 'OMPL', 'SAIL', 'SANVI', 'SWASTIK', 'SYPNL', 'TTL', 'SABBL', 'HFIN', 'RSML', 'SOHL'
    ]  # Example: replace with your desired symbols

    df = pd.DataFrame(floorsheet)
    combined_filename = f"Combined_selected_symbols_{today_str}.xlsx"

    def sanitize_sheet_name(name):
        # Remove or replace invalid characters for Excel sheet names
        return re.sub(r"[\\/*\[\]:?']", "_", str(name))[:31]  # Excel sheet names max 31 chars

    with pd.ExcelWriter(combined_filename, engine='openpyxl') as writer:
        # Sheet 1: All transactions of the day
        all_df = df.copy()
        if 'contractId' in all_df.columns:
            all_df['contractId'] = all_df['contractId'].astype(str)
        all_df.to_excel(writer, sheet_name='All_Transactions', index=False)

        # Sheets for each selected symbol
        for symbol in selected_symbols:
            group = df[df['stockSymbol'] == symbol].copy()
            if group.empty:
                print(f"No transactions found for {symbol}")
                continue
            # Remove 'contractId' column if present
            if 'contractId' in group.columns:
                group = group.drop(columns=['contractId'])
            # Add S.N column starting from 1
            group.insert(0, 'S.N', range(1, len(group) + 1))
            # Add Date column with today's date as a datetime.date object (for Excel date formatting)
            group['Date'] = today_excel_date
            # Retain columns up to and including 'contractAmount', after S.N and Date
            cols = group.columns.tolist()
            if 'contractAmount' in cols:
                idx = cols.index('contractAmount')
                main_cols = ['S.N', 'Date'] + [c for c in cols if c not in ['S.N', 'Date']][:idx-1]
                main_cols += ['contractAmount']
                group = group[main_cols]
            else:
                main_cols = ['S.N', 'Date'] + [c for c in cols if c not in ['S.N', 'Date']][:6]
                group = group[main_cols]
            safe_name = sanitize_sheet_name(symbol)
            group.to_excel(writer, sheet_name=safe_name, index=False)

    # Format the Date column in all symbol sheets as M/D/YYYY
    wb = load_workbook(combined_filename)
    for ws in wb.worksheets:
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col_idx).value == 'Date':
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.number_format = 'm/d/yyyy'
    wb.save(combined_filename)
    print(f"Saved all transactions and selected symbols to '{combined_filename}' with proper formatting.")
except Exception as e:
    print(f"Error saving combined transactions: {e}")